"""End-to-end test on the real Odoo sample.

Validates that the spec contract holds:
- validation produces the expected anomaly count
- the generated .xlsm still contains xl/vbaProject.bin + xl/xmlMaps.xml + xl/tables/
"""
from __future__ import annotations

import datetime as dt
import re
import sys
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app import build_edi_xlsm, load_odoo, validate_and_transform  # noqa: E402

SAMPLE = ROOT / "samples" / "Odoo_template.xlsx"
TEMPLATE = ROOT / "templates" / "EDI_MAROCAINE_XML_GENERATOR.xlsm"
OUT = ROOT / "scripts" / "out_test.xlsm"


def _check(failures: list[str], condition: bool, message: str) -> None:
    status = "PASS" if condition else "FAIL"
    print(f"  {status:<4} {message}")
    if not condition:
        failures.append(message)


def _zip_part_bytes(zf: zipfile.ZipFile, part_name: str) -> bytes | None:
    if part_name not in zf.namelist():
        return None
    return zf.read(part_name)


def _check_identical_part(
    failures: list[str],
    template_zip: zipfile.ZipFile,
    generated_zip: zipfile.ZipFile,
    part_name: str,
    *,
    optional: bool = False,
) -> None:
    template_bytes = _zip_part_bytes(template_zip, part_name)
    generated_bytes = _zip_part_bytes(generated_zip, part_name)

    if template_bytes is None and optional:
        print(f"  SKIP {part_name} absent from template")
        return

    _check(failures, template_bytes is not None, f"{part_name} exists in template")
    _check(failures, generated_bytes is not None, f"{part_name} exists in generated file")

    if template_bytes is None or generated_bytes is None:
        return

    _check(
        failures,
        generated_bytes == template_bytes,
        f"{part_name} is byte-identical to template",
    )


def main() -> int:
    print("=" * 72)
    print("LOADING SAMPLE")
    print("=" * 72)
    df = load_odoo(SAMPLE.read_bytes())
    print(f"  shape = {df.shape}")

    print()
    print("=" * 72)
    print("VALIDATION")
    print("=" * 72)
    df_edi, anomalies = validate_and_transform(df)
    print(f"  valid rows = {len(df_edi)}")
    print(f"  anomalies  = {len(anomalies)}")
    print()
    print("First 15 anomalies:")
    for a in anomalies[:15]:
        print(
            f"  L{a['Ligne Excel']:>3}  "
            f"ref={a['Référence']!r:>18}  "
            f"partner={a['Partenaire']!r:>30}  "
            f"errs={a['Erreurs']}"
        )
    if len(anomalies) > 15:
        print(f"  ...and {len(anomalies) - 15} more")

    if df_edi.empty:
        print("WARN: no valid rows in sample - will still generate empty xlsm to test preservation")

    print()
    print("=" * 72)
    print("GENERATION")
    print("=" * 72)
    header = {
        "raison_sociale": "AitOukhaliTravaux",
        "if": "12345678",
        "annee": 2026,
        "periode": 5,
        "regime": 2,
    }
    xlsm = build_edi_xlsm(df_edi, header, TEMPLATE.read_bytes())
    OUT.write_bytes(xlsm)
    print(f"  wrote {OUT}  ({len(xlsm):,} bytes)")

    expected_last_row = 8 + len(df_edi) + 1
    expected_table_ref = f"A8:M{expected_last_row}"
    expected_filter_ref = f"A8:M{expected_last_row - 1}"
    failures: list[str] = []

    print()
    print("=" * 72)
    print("XLSM PRESERVATION CONTRACT")
    print("=" * 72)
    _check(failures, zipfile.is_zipfile(OUT), "generated file is a valid ZIP container")
    if zipfile.is_zipfile(OUT):
        with zipfile.ZipFile(TEMPLATE) as template_zip, zipfile.ZipFile(OUT) as generated_zip:
            bad_entry = generated_zip.testzip()
            _check(failures, bad_entry is None, "generated ZIP has no corrupt entries")

            _check_identical_part(
                failures, template_zip, generated_zip, "xl/vbaProject.bin"
            )
            _check_identical_part(
                failures, template_zip, generated_zip, "xl/xmlMaps.xml"
            )
            _check_identical_part(
                failures,
                template_zip,
                generated_zip,
                "xl/tables/tableSingleCells1.xml",
                optional=True,
            )
            _check_identical_part(
                failures,
                template_zip,
                generated_zip,
                "xl/drawings/drawing1.xml",
                optional=True,
            )

            template_names = set(template_zip.namelist())
            generated_names = set(generated_zip.namelist())

            drawing_rels = "xl/drawings/_rels/drawing1.xml.rels"
            if drawing_rels in template_names:
                _check(failures, drawing_rels in generated_names, f"{drawing_rels} exists")
                if drawing_rels in generated_names:
                    template_rels = template_zip.read(drawing_rels).decode("utf-8")
                    generated_rels = generated_zip.read(drawing_rels).decode("utf-8")
                    template_image_rels = template_rels.count("/relationships/image")
                    generated_image_rels = generated_rels.count("/relationships/image")
                    _check(
                        failures,
                        generated_image_rels >= template_image_rels,
                        "drawing relationships still include image relationships",
                    )

            for media_part in sorted(n for n in template_names if n.startswith("xl/media/")):
                _check(failures, media_part in generated_names, f"{media_part} exists")
                if media_part in generated_names:
                    _check(
                        failures,
                        generated_zip.read(media_part) == template_zip.read(media_part),
                        f"{media_part} is byte-identical to template",
                    )
                    if drawing_rels in generated_names:
                        generated_rels = generated_zip.read(drawing_rels).decode("utf-8")
                        _check(
                            failures,
                            Path(media_part).name in generated_rels,
                            f"drawing relationships reference {Path(media_part).name}",
                        )

            content_types = generated_zip.read("[Content_Types].xml").decode("utf-8")
            _check(
                failures,
                "application/vnd.ms-excel.sheet.macroEnabled.main+xml" in content_types,
                "[Content_Types].xml keeps macro-enabled workbook content type",
            )

            table_part = "xl/tables/table1.xml"
            _check(failures, table_part in generated_names, "xl/tables/table1.xml exists")
            if table_part in generated_names:
                table_xml = generated_zip.read(table_part).decode("utf-8")
                table_ref = re.search(r"<table\b[^>]*\bref=\"([^\"]+)\"", table_xml)
                filter_ref = re.search(r"<autoFilter\b[^>]*\bref=\"([^\"]+)\"", table_xml)
                _check(failures, table_ref is not None, "table1.xml contains table ref")
                _check(failures, filter_ref is not None, "table1.xml contains autoFilter ref")
                if table_ref is not None:
                    _check(
                        failures,
                        table_ref.group(1) == expected_table_ref,
                        f"table ref is {expected_table_ref}",
                    )
                if filter_ref is not None:
                    _check(
                        failures,
                        filter_ref.group(1) == expected_filter_ref,
                        f"autoFilter ref is {expected_filter_ref}",
                    )

    print()
    print("=" * 72)
    print("CELL-LEVEL SANITY CHECK")
    print("=" * 72)
    from openpyxl import load_workbook  # noqa: PLC0415
    try:
        wb = load_workbook(OUT, keep_vba=True, data_only=False)
    except Exception as exc:  # noqa: BLE001
        _check(failures, False, f"openpyxl can load generated workbook: {exc}")
        wb = None
    else:
        _check(failures, True, "openpyxl can load generated workbook with keep_vba=True")

    if wb is None:
        print()
        print(f"FAIL: {len(failures)} preservation check(s) failed:")
        for failure in failures:
            print(f"  - {failure}")
        return 1

    ws = wb["EDI"]
    print(f"  C2 (raison sociale) = {ws['C2'].value!r}")
    print(f"  C3 (IF)             = {ws['C3'].value!r}")
    print(f"  C4 (année)          = {ws['C4'].value!r}")
    print(f"  C5 (période)        = {ws['C5'].value!r}")
    print(f"  C6 (régime)         = {ws['C6'].value!r}")
    table = ws.tables["Tableau5"]
    print(f"  Tableau5.ref        = {table.ref!r}")
    if table.autoFilter:
        print(f"  Tableau5.filter     = {table.autoFilter.ref!r}")

    _check(failures, table.ref == expected_table_ref, "openpyxl sees expected table ref")
    _check(
        failures,
        table.autoFilter is not None and table.autoFilter.ref == expected_filter_ref,
        "openpyxl sees expected autoFilter ref",
    )

    totals_row = expected_last_row
    expected_totals = {
        f"A{totals_row}": "Total",
        f"D{totals_row}": "=SUBTOTAL(109,Tableau5[M_HT])",
        f"E{totals_row}": "=SUBTOTAL(109,Tableau5[TVA])",
        f"F{totals_row}": "=SUBTOTAL(109,Tableau5[M_TTC])",
        f"M{totals_row}": "=SUBTOTAL(103,Tableau5[DATE_FAC])",
    }
    for coord, expected_value in expected_totals.items():
        _check(
            failures,
            ws[coord].value == expected_value,
            f"{coord} keeps expected total/formula",
        )

    if df_edi.shape[0] > 0:
        r = 9  # first data row
        print(f"  Row {r}:")
        for c in range(1, 14):
            cell = ws.cell(r, c)
            print(f"    col{c}({chr(64+c)}): value={cell.value!r}, fmt={cell.number_format!r}")

        ice_cell = ws.cell(r, 9)
        taux_cell = ws.cell(r, 10)
        dpai_cell = ws.cell(r, 12)
        dfac_cell = ws.cell(r, 13)
        _check(
            failures,
            isinstance(ice_cell.value, str)
            and len(ice_cell.value) == 15
            and ice_cell.number_format == "@",
            "ICE stays text, 15 chars, format @",
        )
        _check(
            failures,
            isinstance(taux_cell.value, (int, float))
            and 0 <= float(taux_cell.value) <= 1
            and taux_cell.number_format == "0%",
            "TAUX stays numeric decimal with 0% format",
        )
        for label, cell in (("DATE_PAIE", dpai_cell), ("DATE_FAC", dfac_cell)):
            _check(
                failures,
                isinstance(cell.value, dt.datetime)
                and cell.number_format == "yyyy-mm-dd",
                f"{label} stays datetime with yyyy-mm-dd format",
            )
    else:
        _check(failures, False, "sample produced at least one row for format checks")

    print()
    if failures:
        print(f"FAIL: {len(failures)} preservation check(s) failed:")
        for failure in failures:
            print(f"  - {failure}")
        return 1
    print("ALL CHECKS PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
