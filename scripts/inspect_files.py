"""One-shot inspector: validate spec assumptions against the real files."""
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "EDI_MAROCAINE_XML_GENERATOR.xlsm"
SAMPLE = ROOT / "samples" / "Odoo_template.xlsx"

print("=" * 70)
print("TEMPLATE EDI:", TEMPLATE.name, f"({TEMPLATE.stat().st_size} bytes)")
print("=" * 70)
wb = openpyxl.load_workbook(TEMPLATE, keep_vba=True)
print("Sheets:", wb.sheetnames)
if "EDI" in wb.sheetnames:
    ws = wb["EDI"]
    print(f"  EDI dimensions: {ws.dimensions}")
    print(f"  Header (C2-C6):")
    for r in range(2, 7):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        print(f"    B{r}={b!r:25}  C{r}={c!r}")
    print(f"  Table header row 8:")
    for col in range(1, 14):
        print(f"    col{col}({chr(64+col)}): {ws.cell(8, col).value!r}")
    print(f"  Tables:")
    for tname in list(ws.tables):
        t = ws.tables[tname]
        print(f"    name={tname!r}, ref={t.ref!r}, displayName={t.displayName!r}")
        if t.autoFilter:
            print(f"      autoFilter.ref={t.autoFilter.ref!r}")
        # Print calculated column formulas if any
        for tcol in t.tableColumns:
            if tcol.calculatedColumnFormula is not None:
                print(f"      {tcol.name!r} calc-formula: {tcol.calculatedColumnFormula.text!r}")

print()
print("=" * 70)
print("ODOO SAMPLE:", SAMPLE.name)
print("=" * 70)
df = pd.read_excel(SAMPLE)
print(f"Shape: {df.shape}")
print(f"Columns ({len(df.columns)}):")
for i, c in enumerate(df.columns):
    print(f"  {chr(65+i)} | {c!r}")
print()
print("Dtypes:")
print(df.dtypes)
print()
print("First 3 rows:")
print(df.head(3).to_string())
print()
print("Sample of 'Lignes de facture/Taxes' values:")
if "Lignes de facture/Taxes" in df.columns:
    print(df["Lignes de facture/Taxes"].value_counts(dropna=False).head(10))
print()
print("Sample of 'Méthode de paiement' values:")
for col in df.columns:
    if "thode" in col.lower() or "paiement" in col.lower():
        print(f"  Column: {col!r}")
        print(df[col].value_counts(dropna=False))
        break
print()
print("Sample of ICE column:")
for col in df.columns:
    if col.upper() == "ICE":
        print(df[col].head(5))
        print("  dtype:", df[col].dtype)
        break
