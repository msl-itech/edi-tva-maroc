"""EDI TVA Maroc - Streamlit generator (MSL-iTECH).

Maps an Odoo vendor-invoice export to the EDI .xlsm template, preserving its
VBA project and XML Map so the user can export the SIMPL-TVA DGI XML from
Excel after download.

Locked behaviour: see SPEC.md (do not deviate without sign-off).
"""
from __future__ import annotations

import base64
import datetime as _dt
import re
import zipfile
from copy import copy
from io import BytesIO
from pathlib import Path
from textwrap import dedent

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
TEMPLATE_PATH = ROOT / "templates" / "EDI_MAROCAINE_XML_GENERATOR.xlsm"
MSL_LOGO_PATH = ROOT / "assets" / "msl-itech-logo.png"
EDI_LOGO_PATH = ROOT / "assets" / "edi-logo.png"

PAYMENT_METHOD_MAP = {
    "ESPECES": 1,
    "CHEQUE": 2,
    "PRELEVEMENT": 3,
    "VIREMENT": 4,
    "LCN": 5,
}

ODOO_REQUIRED_COLS = [
    "Référence",
    "Libellé",
    "Montant hors taxes",
    "Taxe",
    "Total",
    "Lignes de facture/Taxes",
    "IF",
    "Partenaire",
    "ICE",
    "Méthode de paiement",
    "Date de paiement",
    "Date de facturation",
]

TAUX_REGEX = re.compile(r"(\d+(?:[.,]\d+)?)\s*%")
TOLERANCE_DH = 0.05

TABLE_NAME = "Tableau5"
SHEET_NAME = "EDI"
HEADER_ROW = 8        # table header row in the template
DATA_START_ROW = 9    # first data row

# openpyxl drops these parts on save - we re-inject them from the original template.
# Why: the DGI XML Map is in xmlMaps.xml; tableSingleCells1.xml binds C3..C6 to root
# xpaths (identifiantFiscal/annee/periode/regime). Without them the macro export
# produces no XML at all.
TEMPLATE_INJECT_IF_MISSING = (
    "xl/xmlMaps.xml",
    "xl/tables/tableSingleCells1.xml",
)
# openpyxl re-serializes these but loses content - we overwrite with the template's
# binary version. drawing1.xml carries the orange "Générer XML" shape with
# macro="[0]!Export216"; openpyxl re-emits drawing1.xml without <xdr:sp> nodes,
# so the button disappears even though images survive.
TEMPLATE_OVERWRITE_PARTS = (
    "xl/drawings/drawing1.xml",
)
XML_MAPS_CONTENT_TYPE = "application/xml"
SINGLE_CELLS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.tableSingleCells+xml"
)
XML_MAPS_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/xmlMaps"
)
SINGLE_CELLS_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/tableSingleCells"
)

# Default column widths from the EDI template - used as a floor when computing
# auto-fit widths so we never shrink a column below its designer-intended size.
TEMPLATE_COL_WIDTHS = {
    "A": 3.86, "B": 22.0,  "C": 29.86, "D": 13.86, "E": 13.29,
    "F": 12.43, "G": 13.43, "H": 31.71, "I": 17.14, "J": 8.86,
    "K": 11.0, "L": 14.0,  "M": 11.29,
}
MAX_COL_WIDTH = 50.0
DATA_ROW_HEIGHT = 17.25
TOTALS_ROW_HEIGHT = 12.75


# ---------------------------------------------------------------------------
# UI helpers
# ---------------------------------------------------------------------------
def inject_platform_css() -> None:
    """Inject lightweight theme-aware styling for a cleaner platform feel."""
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Geist:wght@300;400;500;600;700&family=Geist+Mono:wght@400;500&family=Fraunces:opsz,wght@9..144,300;9..144,400;9..144,500&display=swap');
        :root {
            --msl-teal: #0B5663;
            --msl-dark: #071E24;
            --msl-accent-teal: #0D9A41;
            --msl-gold: #F4B14C;
            --msl-soft: #F4F1EC;
            --msl-panel: rgba(11, 86, 99, 0.34);
            --msl-border: rgba(244, 241, 236, 0.14);
            --msl-shadow: rgba(0, 0, 0, 0.34);
        }

        .stApp {
            background:
                radial-gradient(circle at 12% 8%, rgba(244, 177, 76, 0.075), transparent 34rem),
                radial-gradient(circle at 82% 2%, rgba(11, 86, 99, 0.23), transparent 38rem),
                linear-gradient(135deg, #082129 0%, #092A31 45%, #071B20 100%);
            color: var(--msl-soft);
            overflow: hidden !important;
        }

        [data-testid="stAppViewContainer"] {
            background:
                linear-gradient(rgba(244, 241, 236, 0.008) 1px, transparent 1px),
                linear-gradient(90deg, rgba(244, 241, 236, 0.008) 1px, transparent 1px);
            background-size: 72px 72px;
        }

        [data-testid="stAppViewContainer"] > section {
            padding: 0 !important;
            max-width: 100% !important;
        }

        section[data-testid="stMain"] > div {
            padding: 0 !important;
        }

        [data-testid="stHeader"] {
            display: none !important;
        }

        [data-testid="stToolbar"] {
            right: 1rem;
        }

        footer {
            display: none !important;
        }

        .block-container {
            padding: 0 !important;
            max-width: 100% !important;
            width: 100% !important;
        }

        .platform-header {
            padding: 0.6rem 1rem;
            margin-bottom: 0.4rem;
            border-radius: 22px;
            border: 1px solid var(--msl-border);
            background:
                linear-gradient(135deg, rgba(244, 241, 236, 0.09), rgba(244, 241, 236, 0.035)),
                rgba(7, 30, 36, 0.78);
            box-shadow: 0 20px 52px rgba(0, 0, 0, 0.28);
            backdrop-filter: blur(14px);
        }

        .platform-brand-row {
            display: grid;
            grid-template-columns: minmax(0, 1fr) auto;
            align-items: center;
            gap: 1rem;
        }

        .platform-brand-main,
        .platform-brand-secondary {
            display: flex;
            align-items: center;
            gap: 0.75rem;
            min-width: 0;
        }

        .platform-brand-secondary {
            justify-content: flex-end;
            padding: 0.56rem 0.74rem;
            border-radius: 16px;
            border: 1px solid rgba(244, 177, 76, 0.24);
            background: rgba(244, 177, 76, 0.08);
        }

        .platform-brand-logo {
            display: block;
            object-fit: contain;
            flex: 0 0 auto;
        }

        .platform-brand-logo--primary {
            width: 36px;
            height: 36px;
            border-radius: 8px;
        }

        .platform-brand-logo--secondary {
            max-width: 72px;
            max-height: 44px;
        }

        .platform-brand-copy {
            display: flex;
            flex-direction: column;
            min-width: 0;
        }

        .platform-brand-label {
            font-size: 0.76rem;
            font-weight: 650;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: rgba(244, 241, 236, 0.62);
            margin-bottom: 0.12rem;
        }

        .platform-brand-name {
            font-size: 1.42rem;
            font-weight: 780;
            line-height: 1.1;
            color: var(--msl-soft);
        }

        .platform-brand-meta {
            font-size: 0.88rem;
            color: rgba(244, 241, 236, 0.66);
            line-height: 1.2;
        }

        .platform-logo-fallback {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            min-width: 2.8rem;
            min-height: 2.8rem;
            padding: 0.35rem 0.5rem;
            border-radius: 12px;
            border: 1px solid color-mix(
                in srgb,
                var(--text-color) 12%,
                transparent
            );
            font-size: 0.76rem;
            font-weight: 750;
            color: var(--msl-soft);
            background: rgba(217, 164, 65, 0.12);
        }

        .platform-eyebrow {
            font-size: 0.78rem;
            font-weight: 600;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: color-mix(in srgb, var(--text-color) 65%, var(--background-color));
            margin-bottom: 0.35rem;
        }

        .platform-title {
            font-size: 1.95rem;
            font-weight: 700;
            line-height: 1.1;
            margin: 0;
        }

        .platform-subtitle {
            font-size: 0.9rem;
            color: rgba(244, 241, 236, 0.72);
            margin-top: 0.3rem;
            max-width: 74ch;
        }

        .platform-chips {
            display: flex;
            flex-wrap: wrap;
            gap: 0.35rem;
            margin-top: 0.35rem;
        }

        .platform-chip {
            display: inline-flex;
            align-items: center;
            gap: 0.4rem;
            padding: 0.34rem 0.62rem;
            border-radius: 999px;
            border: 1px solid rgba(244, 177, 76, 0.24);
            background: rgba(244, 177, 76, 0.09);
            color: rgba(244, 241, 236, 0.86);
            font-size: 0.8rem;
            font-weight: 600;
            line-height: 1;
        }

        .workflow-strip {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 0.45rem;
            margin: 0 0 0.5rem;
        }

        .workflow-step {
            display: flex;
            align-items: center;
            gap: 0.5rem;
            padding: 0.5rem 0.7rem;
            border-radius: 16px;
            border: 1px solid rgba(244, 241, 236, 0.13);
            background: rgba(7, 30, 36, 0.62);
            box-shadow: inset 0 1px 0 rgba(244, 241, 236, 0.06);
            transition: transform 140ms ease, border-color 140ms ease, background 140ms ease;
        }

        .workflow-step:hover {
            transform: translateY(-1px);
            border-color: rgba(244, 177, 76, 0.28);
            background: rgba(11, 86, 99, 0.42);
        }

        .workflow-index {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 1.58rem;
            height: 1.58rem;
            border-radius: 999px;
            background: rgba(244, 177, 76, 0.17);
            color: var(--msl-gold);
            font-size: 0.78rem;
            font-weight: 760;
            flex: 0 0 auto;
        }

        .workflow-text {
            display: flex;
            flex-direction: column;
            gap: 0.05rem;
            min-width: 0;
        }

        .workflow-label {
            font-size: 0.92rem;
            font-weight: 650;
            line-height: 1.15;
            margin: 0;
        }

        .workflow-hint {
            font-size: 0.78rem;
            color: rgba(244, 241, 236, 0.58);
            margin: 0;
        }

        .section-header {
            margin-bottom: 0.6rem;
        }

        .section-kicker {
            font-size: 0.76rem;
            font-weight: 650;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: var(--msl-gold);
            margin-bottom: 0.2rem;
        }

        .section-title {
            font-size: 1.12rem;
            font-weight: 650;
            line-height: 1.25;
            margin: 0;
        }

        .section-note {
            font-size: 0.92rem;
            color: rgba(244, 241, 236, 0.66);
            margin-top: 0.2rem;
        }

        div[data-testid="stForm"] {
            padding: 0;
            border: none;
        }

        div[data-testid="stVerticalBlockBorderWrapper"] {
            border-radius: 18px;
            border: 1px solid rgba(244, 241, 236, 0.145);
            background:
                linear-gradient(135deg, rgba(244, 241, 236, 0.082), rgba(244, 241, 236, 0.028)),
                rgba(7, 30, 36, 0.72);
            box-shadow: 0 18px 48px rgba(0, 0, 0, 0.28);
            backdrop-filter: blur(12px);
            padding: 0.75rem 1rem 1rem;
            margin-bottom: 0.6rem;
        }

        @media (max-width: 720px) {
            .platform-brand-row {
                grid-template-columns: 1fr;
            }

            .platform-brand-secondary {
                justify-content: flex-start;
                width: fit-content;
            }

            .workflow-strip {
                grid-template-columns: 1fr 1fr;
            }
        }

        div[data-testid="stExpander"] {
            border-color: rgba(244, 241, 236, 0.14);
            background: rgba(7, 30, 36, 0.35);
        }

        div[data-testid="stDataFrame"] {
            border-radius: 0.5rem;
            overflow: hidden;
        }

        div[data-testid="stAlert"] {
            border-radius: 14px;
            border: 1px solid rgba(244, 177, 76, 0.18);
            background:
                linear-gradient(135deg, rgba(11, 86, 99, 0.32), rgba(7, 30, 36, 0.5));
            box-shadow: none;
        }

        .stButton > button,
        .stDownloadButton > button {
            min-height: 3.18rem;
            border-radius: 18px;
            border: 1px solid rgba(244, 177, 76, 0.56);
            background:
                linear-gradient(135deg, rgba(244, 177, 76, 0.38), rgba(244, 177, 76, 0.16));
            box-shadow:
                inset 0 1px 0 rgba(244, 241, 236, 0.18),
                0 16px 38px rgba(244, 177, 76, 0.18);
            font-weight: 800;
            letter-spacing: 0.01em;
            transition: transform 140ms ease, box-shadow 140ms ease, border-color 140ms ease;
        }

        .stButton > button:hover,
        .stDownloadButton > button:hover {
            transform: translateY(-1px);
            border-color: rgba(244, 177, 76, 0.66);
            box-shadow:
                inset 0 1px 0 rgba(244, 241, 236, 0.22),
                0 20px 48px rgba(244, 177, 76, 0.25);
        }

        div[data-testid="stTextInput"] input,
        div[data-baseweb="select"] > div {
            min-height: 2.75rem;
            border-radius: 14px;
            border-color: rgba(11, 86, 99, 0.58);
            background-color: rgba(7, 30, 36, 0.26);
            transition: border-color 140ms ease, box-shadow 140ms ease, background 140ms ease;
        }

        div[data-testid="stTextInput"] input:focus,
        div[data-baseweb="select"] > div:focus-within {
            border-color: rgba(244, 177, 76, 0.58);
            box-shadow: 0 0 0 3px rgba(244, 177, 76, 0.1);
        }

@media (prefers-color-scheme: light) {
            .stApp {
                background:
                    radial-gradient(circle at 10% 4%, rgba(244, 177, 76, 0.13), transparent 26rem),
                    radial-gradient(circle at 88% 0%, rgba(11, 86, 99, 0.14), transparent 32rem),
                    linear-gradient(135deg, #F4F1EC 0%, #EEF4F2 45%, #F9F6EF 100%);
                color: #071E24;
            }
            .platform-header,
            div[data-testid="stVerticalBlockBorderWrapper"] {
                border-color: rgba(7, 30, 36, 0.12);
                background:
                    linear-gradient(135deg, rgba(255, 255, 255, 0.72), rgba(255, 255, 255, 0.42)),
                    rgba(244, 241, 236, 0.68);
                color: #071E24;
                box-shadow: 0 18px 46px rgba(7, 30, 36, 0.1);
            }
            .platform-brand-name { color: #071E24; }
            .platform-brand-label,
            .platform-brand-meta,
            .platform-subtitle,
            .workflow-hint,
            .section-note { color: rgba(7, 30, 36, 0.66); }
            .workflow-step {
                border-color: rgba(7, 30, 36, 0.1);
                background: rgba(255, 255, 255, 0.48);
                color: #071E24;
            }
            div[data-testid="stTextInput"] input,
            div[data-baseweb="select"] > div {
                background-color: rgba(255, 255, 255, 0.56);
                border-color: rgba(7, 30, 36, 0.14);
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _logo_img(path: Path, alt: str, class_name: str, fallback: str) -> str:
    if not path.exists():
        return f'<span class="platform-logo-fallback">{fallback}</span>'
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f'<img class="{class_name}" src="data:image/png;base64,{encoded}" alt="{alt}" />'


def render_platform_header() -> None:
    """Render the page header without affecting business flow."""
    msl_logo = _logo_img(
        MSL_LOGO_PATH,
        "MSL-iTECH",
        "platform-brand-logo platform-brand-logo--primary",
        "MSL",
    )
    edi_logo = _logo_img(
        EDI_LOGO_PATH,
        "EDI",
        "platform-brand-logo platform-brand-logo--secondary",
        "EDI",
    )
    st.markdown(
        f"""
        <div class="platform-header">
            <div class="platform-brand-row">
                <div class="platform-brand-main">
                    {msl_logo}
                    <div class="platform-brand-copy">
                        <div class="platform-brand-label">MSL-iTECH</div>
                        <div class="platform-brand-name">EDI TVA Maroc</div>
                <div class="platform-brand-meta">Préparation automatisée de la déclaration TVA</div>
                    </div>
                </div>
                <div class="platform-brand-secondary">
                    {edi_logo}
                    <div class="platform-brand-copy">
                        <div class="platform-brand-label">Format EDI</div>
                        <div class="platform-brand-meta">XML DGI</div>
                    </div>
                </div>
            </div>
            <div class="platform-subtitle">
                Importez votre export Odoo, contrôlez les anomalies bloquantes,
                puis préparez un fichier .xlsm conforme au modèle DGI.
            </div>
            <div class="platform-chips">
                <span class="platform-chip">Contrôles bloquants</span>
                <span class="platform-chip">Export via Excel Windows</span>
                <span class="platform-chip">XML Map conservée</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_workflow_strip() -> None:
    """Render a compact four-step workflow strip."""
    st.markdown(
        """
        <div class="workflow-strip">
            <div class="workflow-step">
                <div class="workflow-index">1</div>
                <div class="workflow-text">
                    <div class="workflow-label">Paramètres</div>
                    <div class="workflow-hint">Header de déclaration</div>
                </div>
            </div>
            <div class="workflow-step">
                <div class="workflow-index">2</div>
                <div class="workflow-text">
                    <div class="workflow-label">Import</div>
                    <div class="workflow-hint">Export Odoo .xlsx</div>
                </div>
            </div>
            <div class="workflow-step">
                <div class="workflow-index">3</div>
                <div class="workflow-text">
                    <div class="workflow-label">Validation</div>
                    <div class="workflow-hint">Contrôle bloquant</div>
                </div>
            </div>
            <div class="workflow-step">
                <div class="workflow-index">4</div>
                <div class="workflow-text">
                    <div class="workflow-label">Génération</div>
                    <div class="workflow-hint">Téléchargement .xlsm</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_section_header(step: str, title: str, note: str | None = None) -> None:
    """Render a consistent section header for the workflow bands."""
    note_html = f'<div class="section-note">{note}</div>' if note else ""
    st.markdown(
        f"""
        <div class="section-header">
            <div class="section-kicker">{step}</div>
            <div class="section-title">{title}</div>
            {note_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_intro_page() -> None:
    st.markdown("""<style>
    header[data-testid="stHeader"]{display:none!important}
    div[data-testid="stToolbar"]{display:none!important}
    footer{display:none!important}
    #MainMenu{display:none!important}
    div[data-testid="stMain"]{padding:0!important}
    div[data-testid="block-container"]{padding:0!important;max-width:100%!important}
    div[data-testid="stVerticalBlock"]{gap:0!important;padding:0!important}
    iframe{border:none!important;display:block!important;width:100%!important}
    .stApp{background:#04161B!important}
    </style>""", unsafe_allow_html=True)

    html = (ROOT / "templates" / "intro_final.html").read_text(encoding="utf-8")
    html = re.sub(r'<button class="cta"[^>]*>.*?</button>', '', html, flags=re.DOTALL)
    components.html(html, height=1000, scrolling=False)

    _, col, _ = st.columns([1, 2, 1])
    with col:
        if st.button(
            "→  Démarrer la préparation EDI  ↵",
            key="cta_nav",
            use_container_width=True,
        ):
            st.session_state.page = "platform"
            st.rerun()

    st.markdown("""<style>
    [data-testid="stButton"] > button {
        background: linear-gradient(180deg, #F1CE7E 0%, #D9A441 55%, #B07F2A 100%) !important;
        color: #1A0F00 !important;
        font-size: 14.5px !important;
        font-weight: 600 !important;
        border: 1px solid rgba(255,221,160,.55) !important;
        border-radius: 12px !important;
        padding: 14px 40px !important;
        box-shadow: 0 18px 36px -14px rgba(217,164,65,.55) !important;
    }
    </style>""", unsafe_allow_html=True)


def load_odoo(uploaded_bytes: bytes) -> pd.DataFrame:
    """Read the Odoo xlsx. Forces str dtype on ICE / IF / Référence to avoid
    pandas turning them into floats (which loses leading zeros)."""
    return pd.read_excel(
        BytesIO(uploaded_bytes),
        dtype={"Référence": str, "IF": str, "ICE": str},
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return False


def _clean_numeric_str(v) -> str:
    """For columns forced to str: drop the trailing '.0' pandas adds when the
    underlying Excel cell was numeric."""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalize_payment(v) -> str | None:
    if _is_blank(v):
        return None
    return str(v).strip().upper()


def parse_taux(raw, taxe: float) -> tuple[float | None, str | None]:
    """Returns (taux, error). Taux is a decimal in [0, 1]."""
    if _is_blank(raw):
        if abs(taxe) < 1e-9:
            return 0.0, None
        return None, "Colonne 'Lignes de facture/Taxes' vide alors que Taxe ≠ 0"
    m = TAUX_REGEX.search(str(raw))
    if not m:
        return None, f"Taux non détectable dans '{raw}'"
    val = float(m.group(1).replace(",", ".")) / 100.0
    return val, None


def normalize_ice(v) -> tuple[str | None, str | None]:
    if _is_blank(v):
        return None, "ICE manquant"
    s = _clean_numeric_str(v)
    if not s.isdigit():
        return None, f"ICE non numérique: '{s}'"
    if len(s) > 15:
        return None, f"ICE > 15 chiffres ({len(s)}): '{s}'"
    return s.zfill(15), None


def normalize_if(v) -> tuple[int | None, str | None]:
    if _is_blank(v):
        return None, "IF manquant"
    s = _clean_numeric_str(v)
    try:
        return int(s), None
    except ValueError:
        return None, f"IF non convertible en entier: '{s}'"


def to_pydate(v) -> _dt.datetime | None:
    if _is_blank(v):
        return None
    if isinstance(v, _dt.datetime):
        return v
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    try:
        ts = pd.to_datetime(v, errors="raise")
        return ts.to_pydatetime() if pd.notna(ts) else None
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Validation + transformation
# ---------------------------------------------------------------------------
def validate_and_transform(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict]]:
    missing_cols = [c for c in ODOO_REQUIRED_COLS if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Colonnes Odoo manquantes: {missing_cols}")

    anomalies: list[dict] = []
    valid_rows: list[dict] = []

    for idx, row in df.iterrows():
        line_num = int(idx) + 2  # Excel row = pandas index + header(1) + 1
        errors: list[str] = []

        ref = row["Référence"]
        if _is_blank(ref):
            errors.append("Référence manquante")
        libelle = row["Libellé"]
        if _is_blank(libelle):
            errors.append("Libellé manquant")
        partner = row["Partenaire"]
        if _is_blank(partner):
            errors.append("Partenaire manquant")

        try:
            ht = float(row["Montant hors taxes"])
            if pd.isna(ht):
                raise ValueError
        except (ValueError, TypeError):
            ht = None
            errors.append("HT manquant ou non numérique")

        try:
            taxe = float(row["Taxe"])
            if pd.isna(taxe):
                raise ValueError
        except (ValueError, TypeError):
            taxe = None
            errors.append("Taxe manquante ou non numérique")

        try:
            total = float(row["Total"])
            if pd.isna(total):
                raise ValueError
        except (ValueError, TypeError):
            total = None
            errors.append("Total manquant ou non numérique")

        if_val, if_err = normalize_if(row["IF"])
        if if_err:
            errors.append(if_err)

        ice_val, ice_err = normalize_ice(row["ICE"])
        if ice_err:
            errors.append(ice_err)

        payment_raw = row["Méthode de paiement"]
        payment_norm = normalize_payment(payment_raw)
        if payment_norm is None:
            errors.append("Méthode de paiement manquante")
            payment_id = None
        elif payment_norm not in PAYMENT_METHOD_MAP:
            errors.append(f"Méthode de paiement inconnue: '{payment_raw}'")
            payment_id = None
        else:
            payment_id = PAYMENT_METHOD_MAP[payment_norm]

        dpai = to_pydate(row["Date de paiement"])
        if dpai is None:
            errors.append("Date de paiement manquante")
        dfac = to_pydate(row["Date de facturation"])
        if dfac is None:
            errors.append("Date de facturation manquante")

        taux_val = None
        if taxe is not None:
            taux_val, taux_err = parse_taux(row["Lignes de facture/Taxes"], taxe)
            if taux_err:
                errors.append(taux_err)

        if (
            ht is not None
            and taxe is not None
            and total is not None
            and abs((ht + taxe) - total) > TOLERANCE_DH
        ):
            errors.append(
                f"Incohérence HT+TVA vs Total: {ht:.2f}+{taxe:.2f}={ht+taxe:.2f} ≠ {total:.2f}"
            )

        if errors:
            anomalies.append({
                "Ligne Excel": line_num,
                "Référence": "(vide)" if _is_blank(ref) else str(ref).strip(),
                "Partenaire": "(vide)" if _is_blank(partner) else str(partner).strip(),
                "Erreurs": " ; ".join(errors),
            })
            continue

        valid_rows.append({
            "OR": len(valid_rows) + 1,
            "FACT_NUM": str(ref).strip(),
            "DESIGNATION": str(libelle).strip(),
            "M_HT": ht,
            "TVA": taxe,
            "M_TTC": total,
            "IF": if_val,
            "LIB_FRSS": str(partner).strip(),
            "ICE_FRS": ice_val,
            "TAUX": taux_val,
            "ID_PAIE": payment_id,
            "DATE_PAIE": dpai,
            "DATE_FAC": dfac,
        })

    df_edi = pd.DataFrame(valid_rows)
    return df_edi, anomalies


# ---------------------------------------------------------------------------
# XML Map preservation (post-openpyxl-save patch)
# ---------------------------------------------------------------------------
def _patch_content_types(xml: str) -> str:
    """Add the two missing Overrides if absent. The output of openpyxl uses
    self-closing tags; we match that style."""
    if "/xl/xmlMaps.xml" not in xml:
        xml = xml.replace(
            "</Types>",
            f'<Override PartName="/xl/xmlMaps.xml" ContentType="{XML_MAPS_CONTENT_TYPE}" />'
            f'<Override PartName="/xl/tables/tableSingleCells1.xml" '
            f'ContentType="{SINGLE_CELLS_CONTENT_TYPE}" /></Types>',
        )
    return xml


def _patch_workbook_rels(xml: str) -> str:
    """Add the xmlMaps relationship to workbook.xml.rels."""
    if 'Target="xmlMaps.xml"' in xml or "xmlMaps.xml" in xml:
        return xml
    # find next free rId
    ids = re.findall(r'Id="rId(\d+)"', xml)
    next_id = max((int(i) for i in ids), default=0) + 1
    new_rel = (
        f'<Relationship Id="rId{next_id}" '
        f'Type="{XML_MAPS_REL_TYPE}" Target="xmlMaps.xml" />'
    )
    return xml.replace("</Relationships>", new_rel + "</Relationships>")


def _patch_sheet1_rels(xml: str) -> str:
    """Add tableSingleCells relationship on the EDI sheet rels file."""
    if "tableSingleCells1.xml" in xml:
        return xml
    ids = re.findall(r'Id="rId(\d+)"', xml)
    next_id = max((int(i) for i in ids), default=0) + 1
    new_rel = (
        f'<Relationship Id="rId{next_id}" Type="{SINGLE_CELLS_REL_TYPE}" '
        f'Target="../tables/tableSingleCells1.xml" />'
    )
    return xml.replace("</Relationships>", new_rel + "</Relationships>")


def _preserve_xml_map(openpyxl_bytes: bytes, template_bytes: bytes) -> bytes:
    """Reconcile what openpyxl produced with the template's binary truth:

    - inject xmlMaps / tableSingleCells if openpyxl dropped them, so the macro
      XML export still finds the DGI schema and root-element bindings;
    - overwrite drawing1.xml with the template's bytes, so the orange shape /
      macro button survives (openpyxl re-serializes drawings but loses <xdr:sp>).
    """
    with zipfile.ZipFile(BytesIO(template_bytes)) as tpl_zip:
        names = set(tpl_zip.namelist())
        inject = {
            name: tpl_zip.read(name)
            for name in TEMPLATE_INJECT_IF_MISSING
            if name in names
        }
        overwrite = {
            name: tpl_zip.read(name)
            for name in TEMPLATE_OVERWRITE_PARTS
            if name in names
        }

    patchers = {
        "[Content_Types].xml": _patch_content_types,
        "xl/_rels/workbook.xml.rels": _patch_workbook_rels,
        "xl/worksheets/_rels/sheet1.xml.rels": _patch_sheet1_rels,
    }

    src = zipfile.ZipFile(BytesIO(openpyxl_bytes))
    out_buf = BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as dst:
        existing = set(src.namelist())
        for name in src.namelist():
            if name in overwrite:
                dst.writestr(name, overwrite[name])
                continue
            data = src.read(name)
            if name in patchers:
                data = patchers[name](data.decode("utf-8")).encode("utf-8")
            dst.writestr(name, data)
        for name, data in inject.items():
            if name not in existing:
                dst.writestr(name, data)
        for name, data in overwrite.items():
            if name not in existing:
                dst.writestr(name, data)
    src.close()
    return out_buf.getvalue()


# ---------------------------------------------------------------------------
# Injection in .xlsm
# ---------------------------------------------------------------------------
def _clone_row_styles(ws, src_row: int) -> list:
    """Snapshot the StyleArray of cells A..M on a given row. Returned so the
    caller can later push the styling onto newly created rows beyond the
    template's original table range."""
    return [copy(ws.cell(row=src_row, column=c)._style) for c in range(1, 14)]


def _apply_row_styles(ws, target_row: int, styles: list) -> None:
    for c in range(1, 14):
        ws.cell(row=target_row, column=c)._style = copy(styles[c - 1])


def _rendered_length(value, number_format: str) -> int:
    """Approximate how many characters Excel will render for the cell, so the
    auto-fit estimator matches what the user sees (not the underlying value)."""
    if value is None:
        return 0
    if isinstance(value, str):
        if value.startswith("="):
            # Formula value - the result is unknown at write time; ignored for sizing.
            return 0
        return len(value)
    if isinstance(value, _dt.date):  # also catches datetime
        return 10  # yyyy-mm-dd
    if isinstance(value, bool):
        return len(str(value))
    if isinstance(value, (int, float)):
        nf = number_format or ""
        if "%" in nf:
            return len(f"{round(value * 100)}%")
        if "#,##0.00" in nf or "# ##0.00" in nf:
            sign = "-" if value < 0 else ""
            return len(sign + f"{abs(value):,.2f}")
        if "#,##0" in nf or "# ##0" in nf:
            sign = "-" if value < 0 else ""
            return len(sign + f"{abs(int(round(value))):,}")
        if isinstance(value, int) or float(value).is_integer():
            return len(str(int(value)))
        return len(f"{value:.2f}")
    return len(str(value))


def _autofit_columns(ws, n: int, last_row: int) -> dict[str, float]:
    """Pick a width per column from header + data + totals content. We clamp at
    the template's designer-intended width on the low end (so we never shrink a
    column the user expects) and at MAX_COL_WIDTH on the high end.

    Totals row formulas (SUBTOTAL) are estimated from the data so the column
    doesn't render ##### when the sum is longer than any individual value."""
    widths: dict[str, float] = {}
    for col_idx in range(1, 14):
        col = get_column_letter(col_idx)
        max_len = 0
        header_val = ws.cell(HEADER_ROW, col_idx).value
        if header_val is not None:
            max_len = max(max_len, len(str(header_val)))

        numeric_values: list[float] = []
        for r in range(DATA_START_ROW, DATA_START_ROW + n):
            cell = ws.cell(r, col_idx)
            v = cell.value
            max_len = max(max_len, _rendered_length(v, cell.number_format))
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numeric_values.append(float(v))

        totals_cell = ws.cell(last_row, col_idx)
        totals_val = totals_cell.value
        if isinstance(totals_val, str) and totals_val.startswith("="):
            if "SUBTOTAL(109" in totals_val and numeric_values:
                estimate = sum(numeric_values)
                max_len = max(max_len, _rendered_length(estimate, totals_cell.number_format))
            elif "SUBTOTAL(103" in totals_val:
                max_len = max(max_len, _rendered_length(n, totals_cell.number_format))
        else:
            max_len = max(max_len, _rendered_length(totals_val, totals_cell.number_format))

        computed = max_len * 1.1 + 2
        floor = TEMPLATE_COL_WIDTHS[col]
        final = max(floor, min(MAX_COL_WIDTH, computed))
        ws.column_dimensions[col].width = final
        widths[col] = round(final, 2)
    return widths


def _apply_row_heights(ws, n: int, last_row: int) -> None:
    """Force a uniform data-row height because openpyxl only carries explicit
    heights for cells the template defined (rows 9..16); rows 17+ otherwise
    fall back to Excel's default 15 and look visibly shorter."""
    for r in range(DATA_START_ROW, DATA_START_ROW + n):
        ws.row_dimensions[r].height = DATA_ROW_HEIGHT
    ws.row_dimensions[last_row].height = TOTALS_ROW_HEIGHT


def build_edi_xlsm(df_edi: pd.DataFrame, header: dict, template_bytes: bytes) -> bytes:
    wb = load_workbook(BytesIO(template_bytes), keep_vba=True)
    ws = wb[SHEET_NAME]

    ws["C2"] = header["raison_sociale"]
    ws["C3"] = header["if"]
    ws["C4"] = int(header["annee"])
    ws["C5"] = int(header["periode"])
    ws["C6"] = int(header["regime"])

    table = ws.tables[TABLE_NAME]
    last_existing = int(table.ref.split(":")[1][1:])   # template totals-row position
    original_last_data_row = last_existing - 1          # template last data row

    # Snapshot the original data-row and totals-row styling. We take row
    # DATA_START_ROW + 2 (a "middle" data row) because rows 9, the first row,
    # and the original last data row have edge-specific borders we don't want
    # to propagate.
    data_row_styles = _clone_row_styles(ws, DATA_START_ROW + 2)
    totals_row_styles = _clone_row_styles(ws, last_existing)

    for r in range(DATA_START_ROW, last_existing + 1):
        for c in range(1, 14):
            ws.cell(row=r, column=c).value = None

    for i, edi_row in enumerate(df_edi.itertuples(index=False)):
        r = DATA_START_ROW + i
        if r > original_last_data_row:
            # New row beyond the template — push the data-row styling so that
            # TableStyleMedium9 striping renders consistently with rows 9..16.
            _apply_row_styles(ws, r, data_row_styles)

        ws.cell(r, 1, int(edi_row.OR))
        ws.cell(r, 2, str(edi_row.FACT_NUM))
        ws.cell(r, 3, str(edi_row.DESIGNATION))
        ws.cell(r, 4, float(edi_row.M_HT))
        ws.cell(r, 5, float(edi_row.TVA))
        ws.cell(r, 6, float(edi_row.M_TTC))
        ws.cell(r, 7, int(edi_row.IF))
        ws.cell(r, 8, str(edi_row.LIB_FRSS))
        ice_cell = ws.cell(r, 9, str(edi_row.ICE_FRS))
        ice_cell.number_format = "@"
        taux_cell = ws.cell(r, 10, float(edi_row.TAUX))
        taux_cell.number_format = "0%"
        ws.cell(r, 11, int(edi_row.ID_PAIE))
        dpai_cell = ws.cell(r, 12, edi_row.DATE_PAIE)
        dpai_cell.number_format = "yyyy-mm-dd"
        dfac_cell = ws.cell(r, 13, edi_row.DATE_FAC)
        dfac_cell.number_format = "yyyy-mm-dd"

    n = len(df_edi)
    last_row = HEADER_ROW + n + 1   # header + n data + 1 totals
    table.ref = f"A{HEADER_ROW}:M{last_row}"
    if table.autoFilter is not None:
        table.autoFilter.ref = f"A{HEADER_ROW}:M{last_row - 1}"

    # Re-emit the totals row at its new position. clear_old_data above wiped
    # the SUBTOTAL formulas that lived on the template's row 17; rewrite them
    # in step with the table column definitions (totalsRowFunction/Label).
    _apply_row_styles(ws, last_row, totals_row_styles)
    ws.cell(last_row, 1, "Total")
    ws.cell(last_row, 4, "=SUBTOTAL(109,Tableau5[M_HT])")
    ws.cell(last_row, 5, "=SUBTOTAL(109,Tableau5[TVA])")
    ws.cell(last_row, 6, "=SUBTOTAL(109,Tableau5[M_TTC])")
    ws.cell(last_row, 13, "=SUBTOTAL(103,Tableau5[DATE_FAC])")

    # If the user has fewer rows than the template default (8), some original
    # data/totals cells now sit outside the new table range. Strip their style
    # so they don't render as orphan blue cells under the new totals row.
    if last_row < last_existing:
        for r in range(last_row + 1, last_existing + 1):
            for c in range(1, 14):
                cell = ws.cell(row=r, column=c)
                cell._style = StyleArray()
                cell.value = None
            if r in ws.row_dimensions:
                ws.row_dimensions[r].height = None

    _apply_row_heights(ws, n, last_row)
    _autofit_columns(ws, n, last_row)

    out = BytesIO()
    wb.save(out)
    return _preserve_xml_map(out.getvalue(), template_bytes)


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
def _safe_name(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "_", s)[:30]


def main() -> None:
    st.set_page_config(page_title="EDI TVA Maroc", page_icon="📑", layout="wide")
    inject_platform_css()

    params = st.query_params
    if params.get("page") == "platform":
        st.query_params.clear()
        st.session_state.page = "platform"
        st.rerun()

    if "page" not in st.session_state:
        st.session_state.page = "intro"
    if st.session_state.page == "intro":
        render_intro_page()
        return

    render_platform_header()
    render_workflow_strip()

    if not TEMPLATE_PATH.exists():
        st.error(f"Template introuvable : {TEMPLATE_PATH}")
        st.stop()

    # 1. Paramètres ----------------------------------------------------------
    with st.container(border=True):
        render_section_header(
            "ÉTAPE 1",
            "Paramètres de la déclaration",
            "Saisissez les informations du header avant d’importer l’export Odoo.",
        )
        c1, c2 = st.columns(2)
        with c1:
            raison_sociale = st.text_input("Raison sociale")
        with c2:
            if_decl = st.text_input("Identifiant fiscal (IF)")

        c3, c4, c5 = st.columns(3)
        with c3:
            regime_label = st.selectbox("Régime", ["Mensuel", "Trimestriel"])
            regime = 1 if regime_label == "Mensuel" else 2
        with c4:
            current_year = _dt.datetime.now().year
            year_options = [current_year - 1, current_year]
            annee = st.selectbox("Année", year_options, index=1)

        period_options = list(range(1, 13)) if regime == 1 else list(range(1, 5))
        period_label = "Période (mois)" if regime == 1 else "Période (trimestre)"
        with c5:
            periode = st.selectbox(
                period_label,
                period_options,
                index=0,
                key=f"periode_{regime}",
            )

        submitted = st.button("✅ Valider paramètres")

    if submitted:
        if not raison_sociale.strip() or not if_decl.strip():
            st.error("Raison sociale et IF sont obligatoires.")
        else:
            st.session_state["header"] = {
                "raison_sociale": raison_sociale.strip(),
                "if": if_decl.strip(),
                "annee": int(annee),
                "periode": int(periode),
                "regime": int(regime),
            }
            st.success("Paramètres validés.")

    header = st.session_state.get("header")
    if header is None:
        st.info("Saisissez et validez les paramètres pour continuer.")
        return

    # 2. Upload --------------------------------------------------------------
    with st.container(border=True):
        render_section_header(
            "ÉTAPE 2",
            "Upload export Odoo (.xlsx)",
            "Glissez-déposez un export Excel Odoo pour lancer la lecture et le contrôle.",
        )
        up = st.file_uploader("Glisse-dépose le fichier Odoo", type=["xlsx"])
        if up is None:
            return

        try:
            df_odoo = load_odoo(up.getvalue())
        except Exception as e:  # noqa: BLE001
            st.error(f"Lecture impossible : {e}")
            return
        st.success(f"{len(df_odoo)} ligne(s) chargée(s).")
        with st.expander("Aperçu des 10 premières lignes"):
            st.dataframe(df_odoo.head(10), use_container_width=True)

    # 3. Validation ----------------------------------------------------------
    with st.container(border=True):
        render_section_header(
            "ÉTAPE 3",
            "Validation et contrôle",
            "Les anomalies bloquantes empêchent la génération du fichier EDI.",
        )
        try:
            df_edi, anomalies = validate_and_transform(df_odoo)
        except ValueError as e:
            st.error(str(e))
            return

        m1, m2, m3 = st.columns(3)
        m1.metric("Lignes en entrée", len(df_odoo))
        m2.metric("Lignes valides", len(df_edi))
        m3.metric("Anomalies", len(anomalies))

        if anomalies:
            st.warning(f"⚠️ Génération bloquée — {len(anomalies)} ligne(s) en anomalie :")
            st.dataframe(pd.DataFrame(anomalies), use_container_width=True)
            return

        if df_edi.empty:
            st.warning("Aucune ligne valide à exporter.")
            return

    # 4. Génération ----------------------------------------------------------
    with st.container(border=True):
        render_section_header(
            "ÉTAPE 4",
            "Génération du fichier EDI .xlsm",
            "Le fichier est préparé pour téléchargement puis ouverture dans Excel Windows.",
        )
        template_bytes = TEMPLATE_PATH.read_bytes()
        if st.button("🚀 Générer", type="primary"):
            with st.spinner("Génération du .xlsm…"):
                xlsm_bytes = build_edi_xlsm(df_edi, header, template_bytes)
            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M")
            fname = (
                f"EDI_TVA_{_safe_name(header['raison_sociale'])}_"
                f"{header['annee']}_M{header['periode']:02d}_{ts}.xlsm"
            )
            st.download_button(
                "⬇️ Télécharger le .xlsm",
                data=xlsm_bytes,
                file_name=fname,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            )
            st.success(f"Fichier prêt : {fname}")


if __name__ == "__main__":
    main()
